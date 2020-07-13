# -*- coding:utf8 -*- #
#-----------------------------------------------------------------------------------
# ProjectName: python3
# FileName: day6
# Author: TangJianjun
# Date: 2020/7/13
# Description:
#-----------------------------------------------------------------------------------
# 递归函数
sum_ls=0
def ls_sum(list):
    for i in list:
        if type(i)==int:
            global sum_ls
            sum_ls +=i
        elif type(i) ==str:
            print(i,'format error')
            pass
        else:
            return ls_sum(i)
    return sum_ls
# 条件代码入口
if __name__ == '__main__':
    # print(ls_sum([1, 2, 3, '2', [1, 2]]))
    class Person_d6:
        name='Tom'
        _age=18
        __num='086***21'
        def spk_name(self):
            print('I\'m %s.'%self.name)
        def _spk_age(self):
            print('I\'m %d years old.'%self._age)
        def __spk_num(self):
            print('My tel-number is %s'%self.__num)
if __name__ == '__main__':
    tom=Person_d6()
    # print(dir(tom))
    # tom._Person__spk_num()#强行调用私有变量与方法
#对csv文件进行操作
dta=open('C:\\test1.csv',mode='r')#r 只读，w 重写，a 追加
# print(dta)
# 使用循环获取数据
# for i in dta:
#     print(i,end='')
# # 对文件操作后，需要关闭文件
# dta.close()
# 导入csv库
import csv
## 以列表形式读取
# dta2=csv.reader(dta)
# for i in dta2:
#     if i[0]!='name':
#         print('%s, %s， %s岁'%(i[0],i[2],i[1]))
# dta.close()
## 以字典形式读取
# dta3=csv.DictReader(dta)
# for i in dta3:
#     print('%s, %s， %s岁'%(i['name'],i['sex'],i['age']))
dta.close()
##使用‘with open as 对象’结构打开文件，不需要单独写关闭语句，结构运行完自动关闭
# with open('C:\\test1.csv',mode='r') as dt:
#     gt=csv.DictReader(dt)
#     for i in gt:
#         print('%s, %s， %s岁' % (i['name'], i['sex'], i['age']))
# 封装与注释
class Get_Csv_Data:
    def csv_file(self,file_path):
        """
        Getting the data from csv file, display with dict
        :param file_path:
        :return:
        """
        with open(file_path,mode='r',encoding='utf8') as f:
            f=csv.DictReader(f)
            for i in f:
                return i
# 写文件
if __name__ == '__main__':
    with open('c:\\test2.csv',mode='w',newline='') as f:
        wf=csv.writer(f)
        wf.writerow(['name','age','sex'])
        data=[['小丽','10','女'],['小李','12','男'],['小力','9','男']]
        for i in data:
            wf.writerow(i)
# 覆盖文件写入
class Write_csv:
    def csv_file(self,filepath,data):
        """
        Writing data into the file,truncating the file first
        :param filepath:
        :param data:
        :return:
        """
        with open(filepath,mode='w',newline='',encoding='utf8') as f:
            f=csv.writer(f)
            for i in data:
                f.writerow(i)
#末尾追加写入
class Append_csv:
    def csv_file(self,filepath,data):
        """
        Appending to the end of the file if it exists
        :param filepath:
        :param data:
        :return:
        """
        with open(filepath,mode='a',newline='',encoding='utf8') as f:
            f=csv.writer(f)
            for i in data:
                f.writerow(i)
# 对excel文件进行操作
# 创建excel文件编辑并保存
from openpyxl import Workbook#导入openpyxl库
wb=Workbook()
ws=wb.active#获取文件当前活动sheet
# if __name__ == '__main__':
    # ws['A1']='hello'
    # ws.append([1,2])
    # wb.save("C:\\test1.xlsx")
class Write_Excel:
    def xlsx_file(self,filepath,data):
        """
        Appending data to te end of excel file, if file is not exist.
        :param filepath:
        :param data:
        :return:
        """
        for i in data:
            ws.append(i)
        wb.save(filepath)
# 读取excel文件数据
import openpyxl
# if __name__ =='__main__':
#     get_wb=openpyxl.load_workbook('c:\\test1.xlsx')#获取工作簿
#     get_sheet=get_wb['Sheet']#获取工作表
# #     get_cell=get_sheet['A1']#获取指定单元格
# #     print(get_cell.row,get_cell.column,get_cell.value)
# #     print(get_cell)
#     for get_tuple in get_sheet:#获取每行数据(<Cell 'Sheet'.A1>, <Cell 'Sheet'.B1>, <Cell 'Sheet'.C1>)
#         for get_cell in get_tuple:#获取每个单元格数据
#             print(get_cell.coordinate, get_cell.value,end=' | ')
#         print()
class Get_Excel_Data:
    def excel_file(self,filepath,sheet_name):
        get_wb = openpyxl.load_workbook(filepath)
        get_sheet = get_wb[sheet_name]
        for get_tuple in get_sheet:
            for get_cell in get_tuple:
                print( get_cell.value,end='')
        print()
#读取MySQL数据：查询
import pymysql# 导入pymysql库
if __name__ == '__main__':
    # 连接数据库
    my_cnt=pymysql.connect(host='192.168.28.64',
                           port=3306,
                           user='root',
                           password='123456',
                           database='test',
                           charset='utf8')
    # 建立游标，逐行读取（遍历数据库里面的数据，便于操作）
    my_data=my_cnt.cursor()
    # 执行SQL语句
    sql1="insert into person(name,age,sex) values('小花',15,'女')"
    sql2='select p.name 姓名 ,p.age 年龄,p.sex 性别 from person p'
    my_data.execute(sql1)
    my_cnt.commit()#提交改变
    my_data.execute(sql2)
    get_data=my_data.fetchall()
    print(get_data)
    my_cnt.close()#关闭连接
pass
